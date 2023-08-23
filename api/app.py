from flask import Flask, request, jsonify
from openpyxl import Workbook, load_workbook
# import face_recognition
from datetime import datetime
import cv2
import numpy as np

app = Flask(__name__)
excel_file = 'attendance_report.xlsx'


# Endpoint for registering a new user
@app.route('/register', methods=['POST'])
def register():
    # Extract the user details and facial image from the request
    matric_no = request.form['MATRIC. NO.']
    name = request.form['NAME']
    department = request.form['DEPARTMENT']
    image = request.files['PASSPORT']
    
    # Save the image to a directory or database for future recognition
    
    # Save registration details in Excel sheet
    save_registration_details(matric_no, name, department)
    
    # Return a success message
    response = {'message': 'User registered successfully'}
    return jsonify(response), 200
@app.route('/login', methods=['POST'])
def login():
    # Extract login details from the request
    matric_no = request.form['MATRIC. NO.']
    name = request.form['NAME']
    image = request.files['PASSPORT']

    # Load the workbook and select the active sheet
    try:
        workbook = load_workbook(excel_file)
    except FileNotFoundError:
        response = {'error': 'Attendance data not found'}
        return jsonify(response), 404

    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[1] == matric_no and row[2] == name:
            # Load the registered face encoding from the sheet
            registered_encoding = row[4]  # Assuming the face encoding is in the 5th column

            # Perform face recognition on the provided image
            img = cv2.imdecode(np.fromstring(image.read(), np.uint8), cv2.IMREAD_UNCHANGED)
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

            detected_encoding = face_rec.compute_face_descriptor(gray)

            # Compare the detected encoding with the registered encoding
            match = np.linalg.norm(np.array(registered_encoding) - np.array(detected_encoding)) < 0.6

            if match:
                response = {'login_status': True}
                return jsonify(response), 200
            else:
                response = {'login_status': False}
                return jsonify(response), 200

    response = {'error': 'User not found'}
    return jsonify(response), 404


#mark attendance
@app.route('/mark', methods=['POST'])
def mark():
    # Extract the user details and facial image from the request
    matric_no = request.form['MATRIC. NO.']
    name = request.form['NAME']
    department = request.form['DEPARTMENT']
    # image = request.files['PASSPORT']
    
    # Save the image to a directory or database for future recognition
    
    # mark attendance in Excel sheet
    mark_attendance(matric_no, name, department)

    # Return a success message
    response = {'message': 'User registered successfully'}
    return jsonify(response), 200

# Endpoint for recognizing a face and marking attendance

@app.route('/recognize', methods=['POST'])
def recognize():
    # Extract the facial image from the request
    image = request.files['PASSPORT']
    
    # Perform facial recognition on the image
    recognized_user = perform_facial_recognition(image)
    
    if recognized_user:
        matric_no = recognized_user['MATRIC. NO.']
        name = recognized_user['NAME']
        department = recognized_user['DEPARTMENT']
        
        # Mark the attendance for the user and update the status
        mark_attendance(matric_no, name, department)
        
        # Return the recognized user details and updated attendance status
        response = {'MATRIC. NO.': matric_no, 'NAME': name, 'DEPARTMENT': department, 'STATUS': recognized_user['STATUS']}
        return jsonify(response), 200
    else:
        # Return an error message if no match is found
        response = {'error': 'Face not recognized'}
        return jsonify(response), 404



# Endpoint for viewing the attendance report
@app.route('/attendance', methods=['GET'])
def view_attendance():
    # Load the attendance data from the Excel sheet
    attendance_data = load_attendance_data()

    # Return the attendance report
    response = {'attendance': attendance_data}
    return jsonify(response), 200


#Other modules

#samve registration
def save_registration_details(matric_no, name, department):
    # Load the workbook or create a new one if it doesn't exist
    try:
        workbook = load_workbook(excel_file)
    except FileNotFoundError:
        workbook = Workbook()

    # Select the active sheet (first sheet by default)
    sheet = workbook.active

    # Append the registration details to the Excel sheet
    row = [sheet.max_row, matric_no, name, department, "", 0, ""]
    sheet.append(row)

    # Save the workbook
    workbook.save(excel_file)

#preform rocognition
def perform_facial_recognition(image):
    # Load the registered face encodings and user details from a file or database
    registered_users = load_registered_users()

    # Load a pre-trained face detection model from OpenCV
    face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

    # Load the image and convert it to grayscale
    img = cv2.imread(image)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # Detect faces in the image
    faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))

    for (x, y, w, h) in faces:
        face_roi = gray[y:y+h, x:x+w]
        
        # Perform face recognition on the detected face region
        recognized_user = recognize_face(face_roi, registered_users)
        if recognized_user:
            return recognized_user

    return None

def recognize_face(face_roi, registered_users):
    # Perform face recognition by comparing face_roi with registered encodings
    # Calculate the face descriptor for the detected face
    detected_encoding = face_rec.compute_face_descriptor(face_roi)

    for user in registered_users:
        registered_encoding = user['encoding']
        
        # Compare the detected encoding with the registered encoding
        match = np.linalg.norm(np.array(registered_encoding) - np.array(detected_encoding)) < 0.6

        if match:
            return user

    return None


def load_registered_users():
    registered_users = []

    # Load the workbook
    try:
        workbook = load_workbook(excel_file)
    except FileNotFoundError:
        return registered_users

    # Select the active sheet (first sheet by default)
    sheet = workbook.active

    # Iterate over the rows in the sheet and extract the registered user data
    for row in sheet.iter_rows(min_row=2, values_only=True):
        user = {
            'MATRIC. NO.': row[1],
            'NAME': row[2],
            'DEPARTMENT': row[3],
            'encoding': [],  # Add the facial encoding for each user
            'STATUS': row[5],
            'TIMESTAMP': row[6],
        }
        registered_users.append(user)

    return registered_users


def mark_attendance(matric_no, name, department):
    # Load the workbook
    try:
        workbook = load_workbook(excel_file)
    except FileNotFoundError:
        print(f"Error: {excel_file} not found.")
        return

    # Select the active sheet (first sheet by default)
    sheet = workbook.active

    # Find the row corresponding to the user based on the MATRIC. NO.
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[1] == matric_no:
            # Increment the attendance status by 1
            status = row[5]
            new_status = status + 1
            
            # Update the STATUS field with the incremented value
            sheet.cell(row=row[0], column=6, value=new_status)
            
            # Update the TIMESTAMP field with the current date and time
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            sheet.cell(row=row[0], column=7, value=timestamp)
            
            # Save the workbook
            workbook.save(excel_file)
            print(f"Attendance marked for {name} ({matric_no}), {department}.")
            return

    print(f"User with MATRIC. NO. {matric_no} not found in the registration details.")
def load_attendance_data():
    attendance_data = []

    # Load the workbook
    try:
        workbook = load_workbook(excel_file)
    except FileNotFoundError:
        print(f"Error: {excel_file} not found.")
        return attendance_data

    # Select the active sheet (first sheet by default)
    sheet = workbook.active

    # Iterate over the rows in the sheet and extract the attendance data
    for row in sheet.iter_rows(min_row=2, values_only=True):
        user_data = {
            'MATRIC. NO.': row[1],
            'NAME': row[2],
            'DEPARTMENT': row[3],
            'STATUS': row[5],
            'TIMESTAMP': row[6],
        }
        attendance_data.append(user_data)

    return attendance_data


if __name__ == '__main__':
    app.run()